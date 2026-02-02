#!/usr/bin/env python3
"""
PUE Data Collector - Appends JSON/CSV data to Excel file
Automatically updates Excel file each time new data is received
"""

import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime


class PUEDataCollector:
    def __init__(self, excel_file='PUE_Datenbank.xlsx', sheet_name='Geräte'):
        self.excel_file = excel_file
        self.sheet_name = sheet_name
        self._initialize_excel()
    
    def _initialize_excel(self):
        """Create Excel file with headers if it doesn't exist"""
        if not Path(self.excel_file).exists():
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            
            # Define headers (matching your GPT configuration)
            headers = [
                'Hersteller', 'Produktkategorie', 'Produktfamilie', 
                'Modellbezeichnung', 'Nennleistung', 'Kühlleistung',
                'Elektrische Aufnahmeleistung', 'Wirkungsgrad_oder_Verlustleistung',
                'COP_EER_IPLV', 'Teillast_25%', 'Teillast_50%', 'Teillast_75%',
                'Teillast_100%', 'Betriebsbedingungen', 'Quelle_Dateiname',
                'Quelle_Seitenzahl', 'Quelle_Zitat', 'Fehlende_Angaben',
                'Verarbeitungsfehler', 'Zeitstempel'
            ]
            
            # Add headers with formatting
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            column_widths = {
                'A': 15, 'B': 20, 'C': 20, 'D': 25, 'E': 15, 'F': 15,
                'G': 20, 'H': 25, 'I': 15, 'J': 12, 'K': 12, 'L': 12,
                'M': 12, 'N': 30, 'O': 20, 'P': 12, 'Q': 50, 'R': 30,
                'S': 30, 'T': 20
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(self.excel_file)
            print(f"✓ Excel-Datei erstellt: {self.excel_file}")
    
    def add_json_data(self, json_data):
        """
        Add JSON data to Excel file
        json_data: Either a JSON string or parsed dict/list
        """
        # Parse JSON if it's a string
        if isinstance(json_data, str):
            try:
                data = json.loads(json_data)
            except json.JSONDecodeError as e:
                print(f"✗ JSON-Parsing-Fehler: {e}")
                return False
        else:
            data = json_data
        
        # Ensure data is a list
        if not isinstance(data, list):
            data = [data]
        
        # Load existing workbook
        wb = load_workbook(self.excel_file)
        ws = wb[self.sheet_name]
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Add timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add each record
        rows_added = 0
        for record in data:
            # Extract Teillastdaten
            teillast = record.get('Teillastdaten', {}) or {}
            
            # Extract Quelle
            quelle = record.get('Quelle', {}) or {}
            
            # Prepare fehlende_angaben as comma-separated string
            fehlende = record.get('Fehlende_Angaben', [])
            fehlende_str = ', '.join(fehlende) if fehlende else ''
            
            # Prepare row data
            row_data = [
                record.get('Hersteller'),
                record.get('Produktkategorie'),
                record.get('Produktfamilie'),
                record.get('Modellbezeichnung'),
                record.get('Nennleistung'),
                record.get('Kühlleistung'),
                record.get('Elektrische Aufnahmeleistung'),
                record.get('Wirkungsgrad_oder_Verlustleistung'),
                record.get('COP_EER_IPLV'),
                teillast.get('25%'),
                teillast.get('50%'),
                teillast.get('75%'),
                teillast.get('100%'),
                record.get('Betriebsbedingungen'),
                quelle.get('Dateiname'),
                quelle.get('Seitenzahl'),
                quelle.get('Zitat'),
                fehlende_str,
                record.get('Verarbeitungsfehler'),
                timestamp
            ]
            
            # Write row
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col, value=value)
            
            next_row += 1
            rows_added += 1
        
        # Save workbook
        wb.save(self.excel_file)
        print(f"✓ {rows_added} Datensätze hinzugefügt zu {self.excel_file}")
        return True
    
    def add_csv_data(self, csv_string):
        """
        Add CSV data to Excel file
        csv_string: CSV formatted string
        """
        # Parse CSV using pandas
        from io import StringIO
        try:
            df = pd.read_csv(StringIO(csv_string))
            
            # Convert to JSON format and use add_json_data
            records = df.to_dict('records')
            return self.add_json_data(records)
        
        except Exception as e:
            print(f"✗ CSV-Parsing-Fehler: {e}")
            return False
    
    def get_summary(self):
        """Get summary of database contents"""
        df = pd.read_excel(self.excel_file, sheet_name=self.sheet_name)
        
        summary = {
            'Gesamtanzahl': len(df),
            'Hersteller': df['Hersteller'].nunique() if 'Hersteller' in df else 0,
            'Produktkategorien': df['Produktkategorie'].nunique() if 'Produktkategorie' in df else 0,
            'Letzte_Aktualisierung': df['Zeitstempel'].max() if 'Zeitstempel' in df else None
        }
        
        return summary


def main():
    """Example usage"""
  _collector = None

def get_collector():
    global _collector
    if _collector is None:
        _collector = PUEDataCollector("PUE_Datenbank.xlsx")
    return _collector
    
    # Example JSON data from your GPT
    example_json = """
    [
      {
        "Hersteller": "Schneider Electric",
        "Produktkategorie": "USV",
        "Produktfamilie": "Galaxy VS",
        "Modellbezeichnung": "Galaxy VS 100kVA",
        "Nennleistung": "100 kVA",
        "Kühlleistung": null,
        "Elektrische Aufnahmeleistung": "102 kW",
        "Wirkungsgrad_oder_Verlustleistung": "96.5%",
        "COP_EER_IPLV": null,
        "Teillastdaten": {
          "25%": "97.0%",
          "50%": "97.5%",
          "75%": "97.0%",
          "100%": "96.5%"
        },
        "Betriebsbedingungen": "25°C, 50% Last",
        "Quelle": {
          "Dateiname": "schneider_galaxy_vs.pdf",
          "Seitenzahl": "5",
          "Zitat": "Efficiency at 50% load: 97.5%"
        },
        "Fehlende_Angaben": ["COP_EER_IPLV", "Kühlleistung"],
        "Verarbeitungsfehler": null
      }
    ]
    """
    
    # Add data
    collector.add_json_data(example_json)
    
    # Get summary
    summary = collector.get_summary()
    print("\n=== Datenbank-Übersicht ===")
    for key, value in summary.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
